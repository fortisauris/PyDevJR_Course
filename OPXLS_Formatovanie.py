import pandas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, LineChart, BarChart, Reference
from copy import copy
import pandas as pd

wb = Workbook()
ws1 = wb.active

# formatovanie bunky
c = ws1.cell(column=1, row=1, value=473643774386834)
c.number_format = '0.00E+00'
c = ws1.cell(column=1, row=2, value=3.14)
c.number_format = '0.000'
c = ws1.cell(column=1, row=3, value=345789123)
c.number_format = '000-000-000'

c.font = Font(name='Arial', size=15, color='FF0000', bold=True, italic=False)

#Objekty Stylov
font = Font(name='Courier', size=12, color='00FF00', bold=True, italic=False)
farba = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
styl_okraja = Side(border_style='thick', color='0000FF')
border = Border(left=styl_okraja, right=styl_okraja, top=styl_okraja, bottom=styl_okraja)
usporiadanie = Alignment(horizontal='center', vertical='center')

# Tu pouzijeme objekty stylov
ws1['A1'].font = font
ws1['B1'].fill = farba
ws1['B2'].border = border

ws1.merge_cells('B4:D5')
ws1['B4'] = 'NADPIS'
ws1["B4"].alignment = usporiadanie
ws1['B4'].fill = farba
# KOPIROVANIE STYLOV
font2 = copy(font)
font2.name = 'Times'
font2.size = 20
font2.color = 'FFFFFF'
ws1['B4'].font = font2

# WORKSHEET 2
ws2 = wb.create_sheet(title='PANDAS')
# DATA PRIPRAVENE PRE PANDAS
data = {'City':['New York', 'London', 'Paris', 'Berlin', 'Praha'],
        'Vzdialenost-mile':[4255,974,825,421,204]}

df = pandas.DataFrame(data)
df['Vzdialenost-km'] = df['Vzdialenost-mile'] * 1.60936
print(df['Vzdialenost-km'])

# nasypeme data do worksheetu
for row in dataframe_to_rows(df, index=False, header=True):
    ws2.append(row)

# KOLACOVY GRAF
pie = PieChart()
pie.style = 13
labels = Reference(ws2, min_col=1, min_row=2, max_row=6)
data = Reference(ws2, min_col=2, min_row=2, max_row=6)
pie.add_data(data, titles_from_data=False)
pie.set_categories(labels)
pie.title = 'Podiel vzdialenosti v milach'
pie.legend.position = 'r'
ws2.add_chart(pie, 'A10')

bar1 = BarChart()
bar1.style = 13
bar1.type = 'col'
labels = Reference(ws2, min_col=1, min_row=2, max_row=6)
data = Reference(ws2, min_col=3, min_row=2, max_row=6)
bar1.add_data(data, titles_from_data=False)
bar1.set_categories(labels)
bar1.title = 'Podiel vzdialenosti v kilometroch'
bar1.legend.position = 'r'
ws2.add_chart(bar1, 'A25')

bar2 = BarChart()
bar2.style = 7
bar2.type = 'bar'
labels = Reference(ws2, min_col=1, min_row=2, max_row=6)
data = Reference(ws2, min_col=3, min_row=2, max_row=6)
bar2.add_data(data, titles_from_data=False)
bar2.set_categories(labels)
bar2.title = 'Podiel vzdialenosti v kilometroch'
bar2.legend.position = 'r'
ws2.add_chart(bar2, 'J10')

lin = LineChart()
lin.style = 5
#lin.type = 'bar'
labels = Reference(ws2, min_col=1, min_row=2, max_row=6)
data = Reference(ws2, min_col=2, min_row=2, max_row=6)
lin.add_data(data, titles_from_data=False)
lin.set_categories(labels)
lin.title = 'Vzdialenosti k BA v milach'
lin.legend.position = 'r'
ws2.add_chart(lin, 'J25')

# Autoformat buniek v ws1
for idx, col in enumerate(ws1.columns, 1):
    print(get_column_letter(idx))
    ws1.column_dimensions[get_column_letter(idx)].auto_size = True

ws2.protection.sheet = True
ws2.protection.enable()
ws2.protection.password = 'BRYNDZOVEHALUSKY'

# Ulozenie do suboru
wb.save('format.xlsx')
print('SUBOR BOL ULOZENY')