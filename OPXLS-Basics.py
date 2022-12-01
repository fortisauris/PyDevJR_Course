# zaciname vzdy importami
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series
import random

# Classy a funkcie idu sem

if __name__ == '__main__':
    wb = Workbook()

    # WORKSHEET 1
    ws1 = wb.active
    ws1.title = 'NAS NAZOV'
    for row in range(1,12):
        ws1.append(range(12))

    # WORKSHEET 2
    ws2 = wb.create_sheet(title='MAIN')
    ws2['A1'] = 53454
    ws2['A2'] = 6543
    ws2['A3'] = 758476
    ws2['A4'] = '=SUM(A1:A3)'

    # WORKSHEET 3
    ws3 = wb.create_sheet(title='DATA')
    for row in range(1,15):
        for col in range(1,15):
            _ = ws3.cell(column=col, row=row, value=random.randint(0,9999))
    print(ws3['B3'].value)
    print(ws3['B3'].number_format)
    ws3.cell(row=3, column=2, value='SIGNAL LOST')

    alphabet = ('A','B','C','D')
    for col in alphabet:
        cell = col+'15'
        formula = '=SUM(%s1:%s14)' % (col, col)
        ws3[cell] = formula

    # ws3.merge_cells('A1:A4')
    # ws3.unmerge_cells('A1:A4')

    ws3.insert_rows(5,2)  # Kde a kolko
    ws3.insert_cols(5,2)
    ws3.delete_rows(5,2)
    ws3.delete_cols(5,2)

    ws3.move_range('A5:B7', rows=0, cols=15, translate=False)

    ws4 = wb.create_sheet(title='FROM PYTHON')

    data = [
        ['MESTO', 'POCET'],
        ['Bratislava', 564],
        ['Banska Bystrica', 798],
        ['Kosice', 987]
    ]
    for r in data:
        ws4.append(r)

    # JEDNODUCHY GRAF
    values = Reference(ws4, min_col=2, min_row=2, max_col=2, max_row=4)
    chart = BarChart()
    chart.add_data(values)
    ws4.add_chart(chart, anchor='B5')


    # ULOZIT DO xlxs
    wb.save(filename='cvicny.xlsx')
    print('SUBOR BOL ULOZENY')



    # wb = load_workbook(filename='cvicny.xlsx')



# TODO Prestavka do 20:10